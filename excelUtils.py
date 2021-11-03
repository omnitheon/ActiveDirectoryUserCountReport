
import openpyxl
from openpyxl.styles import Font
#Creates a workbook to save output to
def generateWorkBook():
    wb = openpyxl.Workbook()
    sheet = wb.active #Assign active sheet to var sheet
    sheet.title #Active Sheet Title
    sheet.title = "RESULTS" #Change the name of the sheet
    #print(wb.sheetnames) #Review results
    sheet = wb["RESULTS"]
    sheet.cell(row=1, column=1).value = "CN"
    sheet.cell(row=1, column=2).value = "Container"
    sheet.cell(row=1, column=3).value = "RECURSIVE_NUM_MEMBERS"
    sheet.cell(row=1, column=4).value = "NUM_MEMBERS"
    sheet.cell(row=1, column=5).value = "logic_check?"
    sheet.cell(row=1, column=6).value = "DOMAIN"
    sheet.cell(row=1, column=7).value = "CN_IS_SUBGROUP?"
    sheet.cell(row=1, column=8).value = "CN_HAS_SUBGROUPS?"
    sheet.cell(row=1, column=9).value = "DN"
    sheet.cell(row=1, column=10).value = "SUBGROUPS_CSV"


    sheet.cell(row=1, column=1).font = Font(bold=True)
    sheet.cell(row=1, column=2).font = Font(bold=True)
    sheet.cell(row=1, column=3).font = Font(bold=True)
    sheet.cell(row=1, column=4).font = Font(bold=True)
    sheet.cell(row=1, column=5).font = Font(bold=True)
    sheet.cell(row=1, column=6).font = Font(bold=True)
    sheet.cell(row=1, column=7).font = Font(bold=True)
    sheet.cell(row=1, column=8).font = Font(bold=True)
    sheet.cell(row=1, column=9).font = Font(bold=True)
    sheet.cell(row=1, column=10).font = Font(bold=True)

    sheet.freeze_panes = 'A2'
    return wb
    
#Helper function to populate workbook objects with data in an iterative manner
def writeToWB(sheet,rowNum,CSV):
    sheet.cell(row=rowNum, column=1).value = CSV[0]
    sheet.cell(row=rowNum, column=2).value = CSV[1]
    sheet.cell(row=rowNum, column=3).value = CSV[2]
    sheet.cell(row=rowNum, column=4).value = CSV[3]
    sheet.cell(row=rowNum, column=5).value = CSV[4]
    sheet.cell(row=rowNum, column=6).value = CSV[5]
    sheet.cell(row=rowNum, column=7).value = CSV[6]
    sheet.cell(row=rowNum, column=8).value = CSV[7]
    sheet.cell(row=rowNum, column=9).value = CSV[8]
    sheet.cell(row=rowNum, column=10).value = CSV[9]